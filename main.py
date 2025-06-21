# Dosyanın en başına bu satırları ekleyin (diğer import'lardan önce)
import os
import sys
script_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, script_dir)
os.chdir(script_dir)


# full_main.py

import curses
import os
import pyodbc
import json
import shutil
import threading
import npyscreen
from datetime import datetime
from decimal import Decimal
from copy import copy
from openpyxl import load_workbook
from logo import logo_text
from settings import user_mapping, database_name, server_name
from datetime import datetime

cancel_query = False
cancel_excel = False

from settings import user_mapping, database_name, server_name, excel_output_folder

COPY_DIR = excel_output_folder
TEMP_DIR = 'temp'

EXCEL_TYPE_PATH = os.path.join(TEMP_DIR, 'excel_type.txt')
MASTER_EXCEL_PATH = os.path.join(TEMP_DIR, 'master-bl.xlsx')
QUERY_SQL_PATH = os.path.join(TEMP_DIR, 'query.sql')
QUERY_RESULT_JSON_PATH = os.path.join(TEMP_DIR, 'query_result.json')
QUERY_RESULT_DATE_PATH = os.path.join(TEMP_DIR, 'query_result_date.txt')
EXCEL_PROGRESS_PATH = os.path.join(TEMP_DIR, 'excel_progress.txt')
EXCEL_CREATED_DATE_PATH = os.path.join(TEMP_DIR, 'excel_created_date.txt')

def read_temp_file(filename, default_value=""):
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except:
        return default_value

def write_temp_file(filename, content):
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)
    except:
        pass


def confirm_box(message):
    """Display a yes/no dialog and return True if the user selects yes."""
    return npyscreen.notify_yes_no(message)


def info_box(message):
    """Display an informational dialog."""
    npyscreen.notify_confirm(message)


# Ana menüye PDF dönüştürme seçeneği eklemek isterseniz:

def convert_last_excel_to_pdf(stdscr):
    h, w = stdscr.getmaxyx()
    stdscr.clear()
    safe_addstr(stdscr, h//2, w//2-30, "Son Excel dosyası PDF'e dönüştürülüyor...")
    stdscr.refresh()
    
    try:
        from excel_to_pdf import convert_excel_to_pdf
        
        # En son oluşturulan Excel dosyasını bul
        today = datetime.now().strftime("%d-%m-%Y")
        backup_filename = f"{today}-BL.xlsx"
        backup_path = os.path.join(COPY_DIR, backup_filename)
        
        if os.path.exists(backup_path):
            pdf_path = convert_excel_to_pdf(backup_path)
            if pdf_path:
                message = f"PDF başarıyla oluşturuldu: {pdf_path}"
            else:
                message = "PDF oluşturulurken bir hata oluştu."
        else:
            message = f"Excel dosyası bulunamadı: {backup_path}"
    except ImportError as e:
        message = f"ImportError: {str(e)}"
    except Exception as e:
        message = f"Genel hata: {str(e)}"
    
    stdscr.clear()
    display_message(stdscr, message + "\nBaşa dönmek için herhangi bir tuşa bas...", h, w)
    stdscr.getch()

# Bu fonksiyonu main fonksiyonunda kullanmak için menüye ekleme:
# menu = ['SQL Sorgusu Çalıştır', 'Excel Dosyasına Yaz', 'Excel -> PDF Dönüştür', 'Çıkış']

def safe_addstr(stdscr, y, x, text):
    h, w = stdscr.getmaxyx()
    if y >= h:
        y = h - 1
    if x < 0:
        x = 0
    if len(text) > w - 2:
        text = text[:w-2]
    try:
        stdscr.addstr(y, x, text)
    except curses.error:
        pass

def draw_logo(stdscr):
    for i, line in enumerate(logo_text):
        safe_addstr(stdscr, i + 1, 2, line)

def draw_static_info(stdscr, h, w):
    info_lines = ["Programming By", "Cuneyt YENER, © 2025", "quanticvision.co.uk"]
    for i, line in enumerate(info_lines):
        safe_addstr(stdscr, h - 5 + i, w - len(line) - 2, line)

def draw_query_result_date(stdscr, w):
    query_date = read_temp_file(QUERY_RESULT_DATE_PATH, "(no date)")
    text = f"Query Result Date: {query_date}"
    safe_addstr(stdscr, 2, w - len(text) - 2, text)

def draw_excel_info(stdscr, h):
    progress = read_temp_file(EXCEL_PROGRESS_PATH, "0%")
    created_date = read_temp_file(EXCEL_CREATED_DATE_PATH, "(no date)")
    excel_type = read_temp_file(EXCEL_TYPE_PATH, "(no type)")  # << burasi eklendi
    text1 = f"Excel Progress: {progress}"
    text2 = f"Excel Created: {created_date}"
    text3 = f"Excel Type: {excel_type}"  # << burasi eklendi
    safe_addstr(stdscr, h - 7, 2, text1)
    safe_addstr(stdscr, h - 6, 2, text2)
    safe_addstr(stdscr, h - 5, 2, text3)  # << burasi eklendi


def clone_row(ws, src_row_idx, tgt_row_idx):
    for col in range(1, 12):
        src_cell = ws.cell(row=src_row_idx, column=col)
        tgt_cell = ws.cell(row=tgt_row_idx, column=col)
        tgt_cell.value = src_cell.value
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.number_format = src_cell.number_format
        tgt_cell.protection = copy(src_cell.protection)
        tgt_cell.alignment = copy(src_cell.alignment)

def execute_sql_query_worker(show_positive_only_value):

    global cancel_query
    try:
        conn = pyodbc.connect(f'DRIVER={{SQL Server}};SERVER={server_name};DATABASE={database_name};UID=sysdba;PWD=masterkey')
        cursor = conn.cursor()
        with open(QUERY_SQL_PATH, 'r', encoding='utf-8') as f:
            sql_query = f.read()
        cursor.execute(sql_query, show_positive_only_value)


        columns = [column[0] for column in cursor.description]
        results = []




        for row in cursor.fetchall():
            if cancel_query:
                conn.close()
                return False, "Sorgu iptal edildi."
            cleaned_row = {}
            for col, val in zip(columns, row):
                if col == "Sales Representative":
                    continue
                if isinstance(val, Decimal):
                    cleaned_row[col] = str(val)
                elif isinstance(val, datetime):
                    cleaned_row[col] = val.strftime("%d/%m/%Y")
                else:
                    cleaned_row[col] = val

            order_no = cleaned_row.get("Order No", "")
            user_code = order_no[:3] if order_no else 'SO-'
            user_name = user_mapping.get(user_code, user_mapping['SO-'])
            cleaned_row["User Name"] = user_name

            results.append(cleaned_row)  # HER cleaned_row eklenmeli (user name olsa da olmasa da)





        with open(QUERY_RESULT_JSON_PATH, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        now = datetime.now()
        write_temp_file(QUERY_RESULT_DATE_PATH, now.strftime("%d/%m/%Y %H:%M:%S"))
        conn.close()
        return True, "Sorgu basarili.\nBasa donmek icin herhangi bir tusa bas..."
    except Exception as e:
        return False, str(e)

def write_to_excel_worker():
    global cancel_excel
    try:
        if not os.path.exists(MASTER_EXCEL_PATH):
            return False, "Master Excel dosyasi bulunamadi."
        today = datetime.now().strftime("%d-%m-%Y")
        backup_filename = f"{today}-BL.xlsx"
        backup_path = os.path.join(COPY_DIR, backup_filename)
        shutil.copy(MASTER_EXCEL_PATH, backup_path)
        workbook = load_workbook(backup_path)
        ws = workbook.active

        # K2 hucresine Liste Tarihi yaz
        today_date = datetime.now().strftime("%d/%m/%Y")
        ws["K2"].value = f" {today_date}"

        # J3 hucresine Excel Type yaz
        excel_type = read_temp_file(EXCEL_TYPE_PATH, "TUM LISTE")
        ws["J3"].value = f" {excel_type}"


        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row[:11]:
                cell.value = None
        with open(QUERY_RESULT_JSON_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        total = len(data)
        for idx, item in enumerate(data, start=1):
            if cancel_excel:
                return False, "Excel olusturma iptal edildi."
            row_idx = idx + 4
            clone_row(ws, 5, row_idx)
            ws[f"A{row_idx}"].value = idx
            ws[f"B{row_idx}"].value = item.get("Customer / Supplier", "")
            type_mapping = {'( CHEQUE )': 'CHEQUE', '( Return )': 'RETURN', 'Cash': 'CASH', 'TRANSFER': 'TRANSFER', 'Bank/Cash': 'BANK/CASH', '( No Payment )': 'NO PAYMENT'}
            ws[f"C{row_idx}"].value = type_mapping.get(item.get("Last Payment Type", "").strip(), 'OTHER')
            balance_value = item.get("Balance", "")
            try:
                balance_number = float(balance_value.replace(",", "")) if balance_value else None
            except:
                balance_number = None
            ws[f"D{row_idx}"].value = balance_number
            ws[f"E{row_idx}"].value = item.get("POSTCODE", "")
            ws[f"F{row_idx}"].value = item.get("CITY", "")


            # Last Invoice Date
            last_invoice_str = item.get("LastInvoiceDate", "")
            try:
                parsed_date = datetime.strptime(last_invoice_str, "%d/%m/%Y")
                ws[f"G{row_idx}"].value = parsed_date
                ws[f"G{row_idx}"].number_format = "DD/MM/YYYY"
            except:
                ws[f"G{row_idx}"].value = last_invoice_str

            # Last Payment Date
            last_payment_str = item.get("LastPaymentDate", "")
            try:
                parsed_payment = datetime.strptime(last_payment_str, "%d/%m/%Y")
                ws[f"I{row_idx}"].value = parsed_payment
                ws[f"I{row_idx}"].number_format = "DD/MM/YYYY"
            except:
                ws[f"I{row_idx}"].value = last_payment_str

            ws[f"H{row_idx}"].value = item.get("GecenZaman2", "")            
            ws[f"J{row_idx}"].value = item.get("GecenZaman1", "")
            #ws[f"K{row_idx}"].value = item.get("Sales Representative", "")
            #ws[f"K{row_idx}"].value = item.get("Order No", "")
            # Istersen L kolonu olarak User Name de yazabilirsin (opsiyonel)
            ws[f"K{row_idx}"].value = item.get("User Name", "")

            progress = int((idx / total) * 100)
            write_temp_file(EXCEL_PROGRESS_PATH, f"{progress}%")
        workbook.save(backup_path)
        now = datetime.now()
        write_temp_file(EXCEL_CREATED_DATE_PATH, now.strftime("%d/%m/%Y %H:%M:%S"))
        return True, "Excel dosyasi basariyla olusturuldu.\nBasa donmek icin herhangi bir tusa bas..."
    except Exception as e:
        return False, str(e)

def display_message(stdscr, message, h, w):
    lines = message.split('\n')
    for idx, line in enumerate(lines):
        safe_addstr(stdscr, h // 2 + idx, w // 2 - len(line) // 2, line)
    stdscr.refresh()

def main(stdscr):
    global cancel_query, cancel_excel
    curses.curs_set(0)
    curses.start_color()
    curses.init_pair(1, curses.COLOR_BLACK, curses.COLOR_WHITE)
    menu = ['SQL Sorgusu Calistir', 'Excel Dosyasina Yaz', 'PDF''e Donustur', 'Cikis']
    current_row = 0
    while True:
        stdscr.clear()
        h, w = stdscr.getmaxyx()
        draw_logo(stdscr)
        draw_query_result_date(stdscr, w)
        draw_excel_info(stdscr, h)
        draw_static_info(stdscr, h, w)
        for idx, row in enumerate(menu):
            x = w // 2 - len(row) // 2
            y = h // 2 - len(menu) // 2 + idx
            if idx == current_row:
                stdscr.attron(curses.color_pair(1))
                safe_addstr(stdscr, y, x, row)
                stdscr.attroff(curses.color_pair(1))
            else:
                safe_addstr(stdscr, y, x, row)
        stdscr.refresh()
        key = stdscr.getch()
        if key == curses.KEY_UP and current_row > 0:
            current_row -= 1
        elif key == curses.KEY_DOWN and current_row < len(menu) - 1:
            current_row += 1
        elif key == curses.KEY_ENTER or key in [10, 13]:
            if current_row == 0:
                stdscr.clear()
                show_positive_only_value = 1 if confirm_box(
                    "Sadece Balance > 0 olanlari gormek istiyor musunuz?") else 0

                if show_positive_only_value == 1:
                    write_temp_file(EXCEL_TYPE_PATH, "STD BALANCE LIST")
                else:
                    write_temp_file(EXCEL_TYPE_PATH, "TUM LISTE")

                stdscr.clear()
                if not confirm_box(
                        "SQL sorgusu calistirilacak. Devam etmek istiyor musunuz?"):
                    continue



                stdscr.clear()
                safe_addstr(stdscr, h//2, w//2-25, "SQL sorgusu calisiyor... Iptal icin 'Q'")
                stdscr.refresh()
                cancel_query = False
                result = {'success': None, 'message': ''}
                thread = threading.Thread(target=lambda: result.update(zip(['success','message'],execute_sql_query_worker(show_positive_only_value))))
                thread.start()
                spinner = ['|','/','-','\\']
                idx = 0
                while thread.is_alive():
                    stdscr.addstr(h//2+2, w//2, spinner[idx%4])
                    stdscr.refresh()
                    stdscr.timeout(200)
                    if stdscr.getch() in [ord('q'), ord('Q')]:
                        cancel_query = True
                    idx +=1
                stdscr.timeout(-1)
                stdscr.clear()
                display_message(stdscr, result['message'], h, w)
                stdscr.getch()
            elif current_row == 1:
                try:
                    if show_positive_only_value is None:
                        raise NameError
                except NameError:
                    stdscr.clear()
                    info_box("Once SQL sorgusu calistirmalisiniz!")
                    continue

                stdscr.clear()
                if not confirm_box(
                        "Excel dosyasina yazilacak. Devam etmek istiyor musunuz?"):
                    continue



                if show_positive_only_value == 1:
                    write_temp_file(EXCEL_TYPE_PATH, "STD BALANCE LIST")
                else:
                    write_temp_file(EXCEL_TYPE_PATH, "TUM LISTE")

                stdscr.clear()
                safe_addstr(stdscr, h//2, w//2-30, "Excel olusturuluyor... Iptal icin 'Q'")
                stdscr.refresh()
                cancel_excel = False
                result = {'success': None, 'message': ''}
                thread = threading.Thread(target=lambda: result.update(zip(['success','message'],write_to_excel_worker())))
                thread.start()
                spinner = ['|','/','-','\\']
                idx = 0
                while thread.is_alive():
                    stdscr.addstr(h//2+2, w//2, spinner[idx%4])
                    stdscr.refresh()
                    stdscr.timeout(200)
                    if stdscr.getch() in [ord('q'), ord('Q')]:
                        cancel_excel = True
                    idx +=1
                stdscr.timeout(-1)
                stdscr.clear()
                display_message(stdscr, result['message'], h, w)
                stdscr.getch()
            elif current_row == 2:  # Excel -> PDF Donustur seçeneği
                convert_last_excel_to_pdf(stdscr)                
            elif current_row == 3:
                stdscr.clear()
                if not confirm_box("Cikmak istiyor musunuz?"):
                    continue




                break

if __name__ == "__main__":
    curses.wrapper(main)
